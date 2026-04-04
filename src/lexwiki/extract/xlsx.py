"""XLSX to markdown extraction using openpyxl."""

from __future__ import annotations

import io
from pathlib import Path


def extract_xlsx(source: Path) -> tuple[str, None]:
    """Convert XLSX to markdown tables.

    Each sheet becomes a ## heading with a pipe-delimited table.
    Returns (markdown_text, None).
    """
    import openpyxl

    blob = source.read_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(blob), data_only=False, read_only=True)
    lines = []
    max_rows = 500
    total_rows = 0

    for sheet in wb.sheetnames:
        if total_rows >= max_rows:
            lines.append(f"\n*[Truncated at {max_rows} rows]*")
            break
        ws = wb[sheet]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(row)
            total_rows += 1
            if total_rows >= max_rows:
                break
        if not rows:
            continue
        lines.append(f"## Sheet: {sheet}")
        header = [str(c) if c is not None else "" for c in rows[0]]
        lines.append("| " + " | ".join(header) + " |")
        lines.append("| " + " | ".join(["---"] * len(header)) + " |")
        for row in rows[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            while len(cells) < len(header):
                cells.append("")
            cells = cells[: len(header)]
            lines.append("| " + " | ".join(cells) + " |")

    wb.close()
    md = "\n".join(lines) if lines else "[Empty spreadsheet]"
    return md, None
